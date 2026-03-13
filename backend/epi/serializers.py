from datetime import date, timedelta, datetime
from django.db import transaction
from django.http import HttpResponse
from django.db.models import Sum, F
from django.contrib.auth import authenticate

from rest_framework.response import Response
from rest_framework import viewsets, permissions, decorators, response, status
from rest_framework.viewsets import ModelViewSet
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.decorators import action

from .models import EPI, EstoqueEPI, Funcionario, MatrizFuncaoEPI, EntregaEPI, TrocaEPI, Usuario
from .serializers import (
    EPISerializer, EstoqueEPISerializer, FuncionarioSerializer,
    MatrizFuncaoEPISerializer, EntregaEPISerializer, TrocaEPISerializer,
    EntregaEPIRelatorioSerializer, UsuarioSerializer
)


class UsuarioViewSet(ModelViewSet):
    queryset = Usuario.objects.all().order_by("nome")
    serializer_class = UsuarioSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]


class EPIViewSet(viewsets.ModelViewSet):
    queryset = EPI.objects.all().order_by("nome")
    serializer_class = EPISerializer
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=True, methods=["get"])
    def lotes(self, request, pk=None):
        qs = EstoqueEPI.objects.filter(epi_id=pk, quantidade__gt=0).values(
            "id", "lote", "quantidade", "nf_numero", "data_compra"
        )
        return response.Response(list(qs))


class EstoqueEPIViewSet(viewsets.ModelViewSet):
    queryset = EstoqueEPI.objects.all().order_by("-data_compra")
    serializer_class = EstoqueEPISerializer
    permission_classes = [permissions.IsAuthenticated]


@decorators.api_view(["GET"])
def alertas_minimo(request):
    abaixo = (
        EstoqueEPI.objects.values("epi_id", "epi__nome", "epi__alerta_estoque_min")
        .annotate(saldo=Sum("quantidade"))
        .filter(saldo__lt=F("epi__alerta_estoque_min"))
        .order_by("epi__nome")
    )
    return response.Response(list(abaixo))


class FuncionarioViewSet(viewsets.ModelViewSet):
    queryset = Funcionario.objects.all().order_by("nome")
    serializer_class = FuncionarioSerializer
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=True, methods=["post"], url_path="excluir_com_senha")
    def excluir_com_senha(self, request, pk=None):
        funcionario = self.get_object()
        admin_password = (
            request.data.get("senha")
            or request.data.get("admin_password")
            or request.data.get("password")
            or ""
        )

        if not admin_password:
            return Response(
                {"detail": "Informe a senha do administrador."},
                status=status.HTTP_400_BAD_REQUEST
            )

        user = request.user
        is_admin = getattr(user, "perfil", None) == "admin" or user.is_superuser or user.is_staff
        if not is_admin:
            return Response(
                {"detail": "Apenas administradores podem excluir funcionários."},
                status=status.HTTP_403_FORBIDDEN
            )

        auth_user = authenticate(username=user.username, password=admin_password)
        if not auth_user:
            return Response(
                {"detail": "Senha do administrador incorreta."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if funcionario.entregas.exists():
            return Response(
                {"detail": "Este funcionário possui histórico de entregas e não pode ser excluído."},
                status=status.HTTP_400_BAD_REQUEST
            )

        nome = funcionario.nome
        funcionario.delete()

        return Response(
            {"detail": f'Funcionário "{nome}" excluído com sucesso.'},
            status=status.HTTP_200_OK
        )


class MatrizFuncaoEPIViewSet(viewsets.ModelViewSet):
    queryset = MatrizFuncaoEPI.objects.all()
    serializer_class = MatrizFuncaoEPISerializer
    permission_classes = [permissions.IsAuthenticated]


class EntregaEPIViewSet(viewsets.ModelViewSet):
    queryset = EntregaEPI.objects.select_related("funcionario", "epi").all().order_by("-data_entrega", "-id")
    serializer_class = EntregaEPISerializer
    permission_classes = [permissions.IsAuthenticated]

    @decorators.action(detail=False, methods=["post"])
    def entregar(self, request):
        try:
            epi_id = int(request.data.get("epi_id"))
            funcionario_id = int(request.data.get("funcionario_id"))
            quantidade = int(request.data.get("quantidade", 1))
            lote_id = int(request.data.get("lote_id"))
            verif_facial_score = float(request.data.get("verif_facial_score", 0))
        except Exception:
            return response.Response({"detail": "Payload inválido."}, status=400)

        try:
            epi = EPI.objects.get(id=epi_id)
            func = Funcionario.objects.get(id=funcionario_id)
            lote = EstoqueEPI.objects.get(id=lote_id, epi_id=epi_id)
        except (EPI.DoesNotExist, Funcionario.DoesNotExist, EstoqueEPI.DoesNotExist):
            return response.Response({"detail": "EPI/Funcionário/Lote inválido."}, status=404)

        if epi.validade_ca and epi.validade_ca < date.today():
            return response.Response({"detail": "CA vencido. Não é permitida a entrega."}, status=400)

        validade = date.today() + timedelta(days=epi.tempo_validade_dias)

        with transaction.atomic():
            if lote.quantidade < quantidade:
                return response.Response({"detail": "Estoque insuficiente no lote."}, status=400)
            lote.quantidade = lote.quantidade - quantidade
            lote.save(update_fields=["quantidade"])

            ent = EntregaEPI.objects.create(
                funcionario=func,
                epi=epi,
                lote=lote.lote,
                quantidade=quantidade,
                data_validade_prevista=validade,
                verif_facial_score=verif_facial_score,
                responsavel=request.user if request.user.is_authenticated else None,
            )

        ser = EntregaEPISerializer(ent)
        return response.Response(ser.data, status=201)

    def _filtrar_entregas(self, request):
        qs = (
            EntregaEPI.objects.select_related("funcionario", "epi")
            .all()
            .order_by("-data_entrega", "-id")
        )

        setor = request.query_params.get("setor")
        categoria = request.query_params.get("categoria")
        numero_ca = request.query_params.get("numero_ca")
        data_de = request.query_params.get("data_de")
        data_ate = request.query_params.get("data_ate")
        funcionario_id = request.query_params.get("funcionario_id")
        epi_id = request.query_params.get("epi_id")

        if setor:
            qs = qs.filter(funcionario__setor__iexact=setor)
        if categoria:
            qs = qs.filter(epi__categoria=categoria)
        if numero_ca:
            qs = qs.filter(epi__numero_ca__icontains=numero_ca)
        if funcionario_id:
            qs = qs.filter(funcionario_id=funcionario_id)
        if epi_id:
            qs = qs.filter(epi_id=epi_id)

        def parse_date(value):
            try:
                return datetime.strptime(value, "%Y-%m-%d").date()
            except Exception:
                return None

        d0 = parse_date(data_de)
        d1 = parse_date(data_ate)
        if d0:
            qs = qs.filter(data_entrega__date__gte=d0)
        if d1:
            qs = qs.filter(data_entrega__date__lte=d1)

        return qs

    @action(detail=False, methods=["get"])
    def relatorio(self, request):
        qs = self._filtrar_entregas(request)
        page = self.paginate_queryset(qs)
        serializer = EntregaEPIRelatorioSerializer(page if page is not None else qs, many=True)
        if page is not None:
            return self.get_paginated_response(serializer.data)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"])
    def exportar(self, request):
        qs = self._filtrar_entregas(request)

        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            return Response(
                {"detail": "Instale openpyxl: pip install openpyxl"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Entregas EPI"

        headers = [
            "Protocolo",
            "Funcionário",
            "Setor",
            "EPI",
            "Número CA",
            "Categoria",
            "Lote",
            "Quantidade",
            "Data Entrega",
            "Validade Até",
        ]
        ws.append(headers)

        for item in qs:
            validade_ate = ""
            try:
                if item.data_entrega and getattr(item.epi, "tempo_validade_dias", None):
                    d0 = item.data_entrega.date() if hasattr(item.data_entrega, "date") else item.data_entrega
                    validade_ate = (d0 + timedelta(days=int(item.epi.tempo_validade_dias))).isoformat()
            except Exception:
                validade_ate = ""

            ws.append([
                item.protocolo or "",
                item.funcionario.nome if item.funcionario else "",
                getattr(item.funcionario, "setor", "") or "",
                item.epi.nome if item.epi else "",
                item.epi.numero_ca if item.epi else "",
                item.epi.categoria if item.epi else "",
                item.lote or "",
                item.quantidade,
                item.data_entrega.strftime("%Y-%m-%d %H:%M") if item.data_entrega else "",
                validade_ate,
            ])

        for idx, _ in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = 22

        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = 'attachment; filename="relatorio_entregas_epi.xlsx"'
        wb.save(resp)
        return resp


class TrocaEPIViewSet(viewsets.ModelViewSet):
    queryset = TrocaEPI.objects.all().order_by("-data_solicitacao")
    serializer_class = TrocaEPISerializer
    permission_classes = [permissions.IsAuthenticated]


@decorators.api_view(["GET"])
def monitor_validade(request):
    dias = int(request.query_params.get("dias", 30))
    alvo = date.today() + timedelta(days=dias)
    qs = EntregaEPI.objects.filter(data_validade_prevista__lte=alvo).values(
        "funcionario__nome", "funcionario__setor", "epi__nome", "data_validade_prevista"
    ).order_by("data_validade_prevista")
    return response.Response(list(qs))


@decorators.api_view(["GET"])
def kpis(request):
    consumo_por_setor = (
        EntregaEPI.objects.values("funcionario__setor")
        .annotate(total=Sum("quantidade"))
        .order_by("-total")[:10]
    )
    top_consumidores = (
        EntregaEPI.objects.values("funcionario__nome")
        .annotate(total=Sum("quantidade"))
        .order_by("-total")[:10]
    )
    proximos_vencimentos = (
        EntregaEPI.objects.filter(data_validade_prevista__lte=date.today() + timedelta(days=30))
        .values("funcionario__nome", "epi__nome", "data_validade_prevista")
        .order_by("data_validade_prevista")[:20]
    )
    return response.Response({
        "consumo_por_setor": list(consumo_por_setor),
        "top_consumidores": list(top_consumidores),
        "proximos_vencimentos": list(proximos_vencimentos),
    })
